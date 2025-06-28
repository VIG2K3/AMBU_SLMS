from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import re
import smtplib
import sys
import random
import barcode
import sqlite3
import os
from sqlite3 import Error
from barcode.writer import ImageWriter
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QComboBox, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
                             QGroupBox, QFormLayout, QDialog, QSplitter, QMessageBox, QFileDialog,
                             QDateEdit, QCalendarWidget, QStyle, QHeaderView, QSizePolicy, QTextEdit)
from PyQt5.QtGui import QColor, QPixmap, QFont
from PyQt5.QtCore import QTimer, Qt, QDate
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU


class DatabaseManager:
    # Initializes the database connection and creates tables.
    def __init__(self, db_file="Product.db"):
        self.db_file = db_file
        self.create_connection()
        self.create_tables()

    # Establishes connection to SQLite database.
    def create_connection(self):
        """Create a database connection to SQLite database"""
        self.conn = None
        try:
            self.conn = sqlite3.connect(self.db_file)
            self.conn.execute("PRAGMA foreign_keys = ON")
        except Error as e:
            print(e)

    # Creates the products table if it doesn't exist.
    def create_tables(self):
        """Create the products table if it doesn't exist"""
        sql_create_products_table = """CREATE TABLE IF NOT EXISTS products (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    category TEXT NOT NULL,
                                    name TEXT NOT NULL,
                                    description TEXT,
                                    quantity INTEGER NOT NULL,
                                    supplier_email TEXT NOT NULL,
                                    barcode TEXT UNIQUE,
                                    test_date TEXT,
                                    created_date TEXT DEFAULT CURRENT_TIMESTAMP,
                                    status TEXT NOT NULL DEFAULT 'Pending',
                                    creator_type TEXT NOT NULL,
                                    creator_username TEXT NOT NULL
                                );"""

        try:
            c = self.conn.cursor()
            c.execute(sql_create_products_table)
            self.conn.commit()
        except Error as e:
            print(e)
    
    # clean and normalize email such like seperare with comma 
    def _process_emails(self, email_string):
        """Helper method to clean and normalize email strings"""
        if not email_string:
            return None

        # Split, strip, and filter empty strings
        emails = [e.strip() for e in email_string.split(',') if e.strip()]

        # Return as comma-separated string if emails exist
        return ', '.join(emails) if emails else None

    # Inserts a new product into the database.
    def add_product(self, product):
        """Add a new product to the products table"""
        sql = '''INSERT INTO products(category, name, description, quantity, supplier_email, barcode, test_date, creator_type, creator_username)
                 VALUES(?,?,?,?,?,?,?,?,?)'''
        try:
            c = self.conn.cursor()
            # Add 'admin' as the creator_type
            product_with_creator = product + ('admin', 'admin')
            c.execute(sql, product_with_creator)
            self.conn.commit()
            return c.lastrowid
        except Error as e:
            print(e)
            return None

    # Updates an existing product in the database.
    def update_product(self, product):
        """Update an existing product"""
        sql = '''UPDATE products
                 SET category = ?, name = ?, description = ?, quantity = ?, supplier_email = ?, barcode = ?, test_date = ?
                 WHERE id = ?'''
        try:
            c = self.conn.cursor()
            c.execute(sql, product)
            self.conn.commit()
            return c.rowcount
        except Error as e:
            print(e)
            return None

    # Deletes a product by its ID.
    def delete_product(self, product_id):
        """Delete a product by product id"""
        sql = 'DELETE FROM products WHERE id = ?'
        try:
            c = self.conn.cursor()
            c.execute(sql, (product_id,))
            self.conn.commit()
            return c.rowcount
        except Error as e:
            print(e)
            return None

    # Retrieves all products from the database.
    def get_all_products(self):
        """Query all products from the database"""
        try:
            c = self.conn.cursor()
            c.execute("SELECT * FROM products ORDER BY id")
            return c.fetchall()
        except Error as e:
            print(e)
            return []

    # Searches products based on type ID, Name, Category and search term.
    def search_products(self, search_type, search_term):
        """Search products based on type and term"""
        try:
            c = self.conn.cursor()

            if search_type == "ID":
                c.execute("SELECT * FROM products WHERE id = ?", (search_term,))
            elif search_type == "Name":
                c.execute("SELECT * FROM products WHERE name LIKE ?", (f'%{search_term}%',))
            elif search_type == "Category":
                c.execute("SELECT * FROM products WHERE category LIKE ?", (f'%{search_term}%',))
            elif search_type == "Description":
                c.execute("SELECT * FROM products WHERE description LIKE ?", (f'%{search_term}%',))
            else:
                return []

            return c.fetchall()
        except Error as e:
            print(e)
            return []


class DatePickerDialog(QDialog):

    # Initializes the dialog with calendar widget.
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Date")
        self.setWindowModality(Qt.ApplicationModal)
        self.setFixedSize(400, 300)

        self.setStyleSheet(
            """QAbstractItemView:enabled {selection-background-color: #ff0000; selection-color: black;}""")

        layout = QVBoxLayout()
        self.setLayout(layout)

        # Create calendar widget
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setMinimumDate(QDate.currentDate())
        self.calendar.setMaximumDate(QDate(2100, 12, 31))
        self.calendar.setSelectedDate(QDate.currentDate())
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)

        # Buttons
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("Cancel")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addWidget(self.calendar)
        layout.addLayout(button_layout)

        # Connect signals
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

    # Returns the selected date from the calendar.
    def get_selected_date(self):
        return self.calendar.selectedDate()


class DateRangePickerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Date Range")
        self.setWindowModality(Qt.ApplicationModal)
        self.setFixedSize(800, 400)
        self.setStyleSheet(
            """QAbstractItemView:enabled {selection-background-color: #ff0000; selection-color: black;}""")
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Create calendar widgets
        calendar_container = QWidget()
        calendar_layout = QHBoxLayout(calendar_container)
        calendar_layout.setContentsMargins(0, 0, 0, 0)

        # From calendar
        from_group = QGroupBox("From")
        from_layout = QVBoxLayout()
        self.from_calendar = QCalendarWidget()
        self.from_calendar.setGridVisible(True)
        self.from_calendar.setMinimumDate(QDate(1900, 1, 1))
        self.from_calendar.setMaximumDate(QDate(2100, 12, 31))
        self.from_calendar.setSelectedDate(QDate.currentDate())
        self.from_calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)  # This removes the week numbers
        from_layout.addWidget(self.from_calendar)
        from_group.setLayout(from_layout)

        # To calendar
        to_group = QGroupBox("To")
        to_layout = QVBoxLayout()
        self.to_calendar = QCalendarWidget()
        self.to_calendar.setGridVisible(True)
        self.to_calendar.setMinimumDate(QDate(1900, 1, 1))
        self.to_calendar.setMaximumDate(QDate(2100, 12, 31))
        self.to_calendar.setSelectedDate(QDate.currentDate())
        self.to_calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)  # This removes the week numbers
        to_layout.addWidget(self.to_calendar)
        to_group.setLayout(to_layout)
        calendar_layout.addWidget(from_group)
        calendar_layout.addWidget(to_group)

        # Buttons
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("Cancel")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addWidget(calendar_container)
        layout.addLayout(button_layout)

        # Connect signals
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

    # Returns a tuple of selected from/to dates.
    def get_selected_dates(self):
        return (self.from_calendar.selectedDate(), self.to_calendar.selectedDate())


class BarcodeGenerator:
    # Initializes barcode generator and creates directory for barcode images.
    def __init__(self):
        self.data = {}
        self.generated_barcodes = []
        self.used_numbers = set()
        # Create BarcodeImages directory if it doesn't exist
        self.barcode_dir = "AdminBarcodeImages"
        os.makedirs(self.barcode_dir, exist_ok=True)

    # Generates a unique 12-digit number for barcodes.
    def generate_unique_number(self):
        while True:
            num = "".join(random.choices("0123456789", k=12))
            if num not in self.used_numbers:
                self.used_numbers.add(num)
                return num

    # Creates a barcode image for a product.
    def generate_barcode(self, product_name):
        if not product_name:
            return None, None

        barcode_format = barcode.get_barcode_class('upc')
        barcode_number = self.generate_unique_number()
        generated = barcode_format(barcode_number, writer=ImageWriter())

        # Create sanitized filename and full path
        safe_name = "".join(c if c.isalnum() else "_" for c in product_name)
        filename = f"{safe_name}_{barcode_number}.png"
        full_path = os.path.join(self.barcode_dir, filename)

        # Save to the BarcodeImages folder only
        generated.save(os.path.join(self.barcode_dir, f"{safe_name}_{barcode_number}"))
        self.generated_barcodes.append(full_path)
        self.data[barcode_number] = [product_name, full_path]

        return barcode_number, full_path


class BarcodePopup(QDialog):

    # Initializes the popup with barcode image.
    def __init__(self, barcode_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Barcode")
        self.setWindowModality(Qt.ApplicationModal)
        layout = QVBoxLayout()
        self.setLayout(layout)
        self.barcode_label = QLabel()
        self.barcode_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.barcode_label)

        if barcode_path:
            try:
                # Check if file exists directly or in fallback folder
                if not os.path.exists(barcode_path):
                    fallback_path = os.path.join("BarcodeImages", os.path.basename(barcode_path))
                    if os.path.exists(fallback_path):
                        barcode_path = fallback_path
                    else:
                        raise FileNotFoundError(f"Barcode image not found at: {barcode_path} or fallback")

                pixmap = QPixmap(barcode_path)
                if pixmap.isNull():
                    raise ValueError("Failed to load barcode image")

                self.barcode_label.setPixmap(pixmap.scaled(400, 200, Qt.KeepAspectRatio))
            except Exception as e:
                self.barcode_label.setText(f"Error loading barcode:\n{str(e)}")
        else:
            self.barcode_label.setText("No barcode path provided")

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)

#Handle email notification
class EmailNotifier:
    def __init__(self):
        self.sender_email = "muhdniys@gmail.com"  # Your Gmail
        self.app_password = "pcryhhyksjnwgxuo"  # Your 16-char app password

    def send_email(self, recipient_emails, subject, body, attachment_path=None):
        """Send email to one or more recipients"""
        if not recipient_emails:
            return False

        try:
            msg = MIMEMultipart()
            msg["From"] = self.sender_email
            msg["To"] = ", ".join(recipient_emails)
            msg["Subject"] = subject

            # Attach the body text
            msg.attach(MIMEText(body, "plain"))

            # Attach barcode image if provided
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, 'rb') as f:
                    img_data = f.read()
                    img = MIMEImage(img_data, name=os.path.basename(attachment_path))
                    msg.attach(img)

            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(self.sender_email, self.app_password)
            server.send_message(msg)
            server.quit()
            return True
        except Exception as e:
            print(f"Error sending email: {str(e)}")
            return False


class ProductManager(QMainWindow):
    def __init__(self, on_data_changed=None, on_chart_refresh=None):
        super().__init__()
        self.on_data_changed = on_data_changed
        self.on_chart_refresh = on_chart_refresh
        self.setStyleSheet("background-color: #d9d9d9;")
        self.active_timers = []
        self.barcode_gen = BarcodeGenerator()
        self.db = DatabaseManager()
        self.email_notifier = EmailNotifier()
        self.setWindowTitle("Product Manager")
        self.setGeometry(100, 100, 1200, 800)
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)

        # Create search group (top section)
        self.create_search_group()
        self.main_layout.addWidget(self.search_group)

        # Create splitter for table and product details
        self.splitter = QSplitter(Qt.Vertical)
        self.main_layout.addWidget(self.splitter)

        # Create table widget
        self.create_table()
        table_container = QWidget()
        table_container.setLayout(QVBoxLayout())
        table_container.layout().addWidget(self.table)
        table_container.layout().setContentsMargins(0, 0, 0, 0)
        self.splitter.addWidget(table_container)

        # Create product details form
        self.create_product_form()
        form_container = QWidget()
        form_container.setLayout(QVBoxLayout())
        form_container.layout().addWidget(self.product_details_group)
        form_container.layout().setContentsMargins(0, 0, 0, 0)
        self.splitter.addWidget(form_container)

        # Set initial sizes table takes 70%, form takes 30%
        self.splitter.setSizes([int(self.height() * 0.7), int(self.height() * 0.3)])

        # Create buttons (bottom section)
        self.create_buttons()
        self.main_layout.addWidget(self.buttons_container)
        self.load_products_from_db()
        if self.on_data_changed:
            self.on_data_changed()
        if self.on_chart_refresh:
            self.on_chart_refresh()

    # Loads all products from database into the table.
    def load_products_from_db(self):
        self.table.setRowCount(0)
        products = self.db.get_all_products()

        for product in products:
            pid, category, name, description, qty, supplier_email, barcode, test_date, created, status, creator_type, creator_username = product
            self.add_table_row(pid, category, name, description, qty, supplier_email, barcode, test_date, created,
                               status, creator_username)

    # Creates the search section of the UI.
    def create_search_group(self):
        self.search_group = QGroupBox("SEARCH PRODUCTS")
        self.search_group.setStyleSheet("""
            QGroupBox {background-color: #d9d9d9; font-size: 14px; font-weight: bold; border: 1px solid #ccc; border-radius: 5px; margin-top: 10px; padding-top: 15px;}
            QGroupBox::title {subcontrol-origin: margin; left: 10px; padding: 0 3px;}""")

        vbox = QVBoxLayout()
        vbox.setContentsMargins(20, 20, 20, 20)

        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(15)

        search_layout.addStretch()

        self.search_combo = QComboBox()
        self.search_combo.addItems(
            ["Select", "ID", "Name", "Category", "Description", "Status: Approved", "Status: Rejected",
             "Status: Pending"])
        self.search_combo.setStyleSheet("""
            QComboBox {background-color: #d9d9d9; padding: 8px; border: 1px solid #ccc; border-radius: 3px; min-width: 100px; background-color: white;}""")
        search_layout.addWidget(self.search_combo)

        search_input_container = QWidget()
        search_input_layout = QHBoxLayout(search_input_container)
        search_input_layout.setContentsMargins(0, 0, 0, 0)
        search_input_layout.setSpacing(5)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search...")
        self.search_input.setStyleSheet("""
            QLineEdit {padding: 8px; border: 1px solid #ccc; border-radius: 3px; min-width: 300px; background-color: white; font-size: 18px;}""")

        # Add calendar button next to search input
        self.calendar_button = QPushButton()
        self.calendar_button.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_FileDialogContentsView')))
        self.calendar_button.setFixedSize(30, 30)
        self.calendar_button.setStyleSheet("""
            QPushButton {border: 1px solid #ccc; border-radius: 3px; background-color: #f0f0f0;}
            QPushButton:hover {background-color: #e0e0e0;}""")
        self.calendar_button.setToolTip("Filter by creation date range")
        self.calendar_button.clicked.connect(self.show_search_date_range_picker)

        search_input_layout.addWidget(self.search_input)
        search_input_layout.addWidget(self.calendar_button)
        search_layout.addWidget(search_input_container)

        self.search_button = QPushButton("Search")
        self.search_button.setStyleSheet("""
            QPushButton {background-color: #b60338; color: white; border: 1px solid #ccc; padding: 8px; font-weight: bold; min-width: 100px;}
            QPushButton:hover {background-color: #f31659;}
            QPushButton:pressed {background-color: #ff4757;}""")
        search_layout.addWidget(self.search_button)

        self.show_all_button = QPushButton("Show All")
        self.show_all_button.setStyleSheet("""
            QPushButton {background-color: #b60338; color: white; border: 1px solid #ccc; padding: 8px; font-weight: bold; min-width: 100px;}
            QPushButton:hover {background-color: #f31659;}
            QPushButton:pressed {background-color: #ff4757;}""")

        search_layout.addWidget(self.show_all_button)
        search_layout.addStretch()

        vbox.addWidget(search_container)
        self.search_group.setLayout(vbox)

        self.search_button.clicked.connect(self.search_products)
        self.show_all_button.clicked.connect(self.show_all_products)

    # Shows date picker for test date selection.
    def show_date_picker(self):
        dialog = DatePickerDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            selected_date = dialog.get_selected_date()
            self.expiry_date_input.setText(selected_date.toString("dd-MM-yyyy"))

    # Shows date range picker for search filtering.
    def show_search_date_range_picker(self):
        dialog = DateRangePickerDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            from_date, to_date = dialog.get_selected_dates()

            # Display the selected range in search field
            date_range = f"{from_date.toString('dd-MM-yyyy')} to {to_date.toString('dd-MM-yyyy')}"
            self.search_input.setText(date_range)

            # Filter the table directly
            self.filter_table_by_date(from_date.toString('dd-MM-yyyy'), to_date.toString('dd-MM-yyyy'))

    # Validates daate for the range 
    def filter_table_by_date(self, from_date, to_date):
        """Filter the table view by created date range"""
        from_date = datetime.strptime(from_date, "%d-%m-%Y")
        to_date = datetime.strptime(to_date, "%d-%m-%Y")

        for row in range(self.table.rowCount()):
            created_item = self.table.item(row, 8) 
            if created_item:
                try:
                    created_date = datetime.strptime(created_item.text(), "%d-%m-%Y")
                    show_row = from_date <= created_date <= to_date
                    self.table.setRowHidden(row, not show_row)
                except ValueError:
                    # If date format is invalid, hide the row
                    self.table.setRowHidden(row, True)

    # Validates date format (DD-MM-YYYY).
    def validate_date(self, date_str):
        if not date_str.strip():
            self.show_message("Error", "Test date cannot be blank", QMessageBox.Warning)
            return False

        if any(c.isalpha() for c in date_str):
            self.show_message("Error", "Test date cannot contain letters", QMessageBox.Warning)
            return False

        if not re.fullmatch(r'\d{2}-\d{2}-\d{4}', date_str):
            self.show_message("Error", "Date must be in DD-MM-YYYY format (e.g. 31-12-2023)", QMessageBox.Warning)
            return False

        try:
            day, month, year = map(int, date_str.split('-'))
            datetime.strptime(date_str, "%d-%m-%Y")  # Will raise ValueError for invalid dates
            return True

        except ValueError:
            self.show_message("Error", "Please enter a valid calendar date", QMessageBox.Warning)
            return False

    # Validates email format.
    def validate_email(self, email):
        """Validate one or more comma-separated emails"""
        if not email.strip():
            return False

        emails = [e.strip() for e in email.split(',') if e.strip()]

        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

        for email in emails:
            if not re.match(pattern, email):
                return False

        local_part, domain_part = email.split('@')

        if not local_part or len(local_part) > 64:
            return False

        if not domain_part or len(domain_part) > 255:
            return False

        if '..' in local_part or '..' in domain_part:
            return False

        if '.' not in domain_part:
            return False

        tld = domain_part.split('.')[-1]
        if len(tld) < 2:
            return False

        return True

    # Shows a message box.
    def show_message(self, title, message, icon=QMessageBox.Information):
        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setIcon(icon)
        msg.exec_()

    # Creates the product details form.
    def create_product_form(self):
        self.product_details_group = QGroupBox("PRODUCT DETAILS")
        self.product_details_group.setStyleSheet("""
            QGroupBox {background-color: #d9d9d9; font-size: 14px; font-weight: bold; border: 1px solid #ccc; border-radius: 5px; margin-top: 10px; padding-top: 15px;}
            QGroupBox::title {subcontrol-origin: margin; left: 10px; padding: 0 3px;}""")

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignCenter)
        main_layout.setContentsMargins(20, 20, 20, 20)

        form_container = QWidget()
        form_container.setMaximumWidth(900)
        form_layout = QHBoxLayout()
        form_layout.setContentsMargins(20, 10, 20, 10)
        form_layout.setSpacing(30)

        left_column = QFormLayout()
        left_column.setFormAlignment(Qt.AlignLeft)
        left_column.setLabelAlignment(Qt.AlignLeft)
        left_column.setSpacing(15)
        left_column.setContentsMargins(10, 10, 10, 10)

        right_column = QFormLayout()
        right_column.setFormAlignment(Qt.AlignLeft)
        right_column.setLabelAlignment(Qt.AlignLeft)
        right_column.setSpacing(15)
        right_column.setContentsMargins(10, 10, 10, 10)

        label_font = QFont()
        label_font.setBold(True)

        category_label = QLabel("PRODUCT CATEGORY:")
        category_label.setFont(label_font)
        self.category_input = QLineEdit()
        self.category_input.setPlaceholderText("Enter product category")
        self.category_input.setMinimumWidth(250)
        self.category_input.setStyleSheet("""
            QLineEdit {padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: white; font-size: 18px;}""")

        name_label = QLabel("PRODUCT NAME:")
        name_label.setFont(label_font)
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Enter product name")
        self.name_input.setMinimumWidth(250)
        self.name_input.setStyleSheet("""
            QLineEdit {padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: white; font-size: 18px;}""")

        description_label = QLabel("DESCRIPTION:")
        description_label.setFont(label_font)
        self.description_input = QTextEdit()
        self.description_input.setPlaceholderText("Enter product description")
        self.description_input.setMaximumHeight(100)
        self.description_input.setStyleSheet("""
            QTextEdit {padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: white; font-size: 18px;}""")

        quantity_label = QLabel("QUANTITY:")
        quantity_label.setFont(label_font)
        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("Enter quantity")
        self.quantity_input.setMinimumWidth(100)
        self.quantity_input.setStyleSheet("""
            QLineEdit {padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: white; font-size: 18px;}""")

        expiry_label = QLabel("TEST DATE:")
        expiry_label.setFont(label_font)
        expiry_layout = QHBoxLayout()
        self.expiry_date_input = QLineEdit()
        self.expiry_date_input.setPlaceholderText("Test Date (DD-MM-YYYY)")
        self.expiry_date_input.setMinimumWidth(120)
        self.expiry_date_input.setStyleSheet("""
            QLineEdit {padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: white;}""")

        self.calendar_button = QPushButton()
        self.calendar_button.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_FileDialogContentsView')))
        self.calendar_button.setFixedSize(30, 30)
        self.calendar_button.setStyleSheet("""
            QPushButton {border: 1px solid #ccc; border-radius: 3px; background-color: #f0f0f0;}
            QPushButton:hover {background-color: #e0e0e0;}""")

        self.calendar_button.clicked.connect(self.show_date_picker)
        expiry_layout.addWidget(self.expiry_date_input)
        expiry_layout.addWidget(self.calendar_button)
        expiry_layout.setSpacing(5)

        supplier_email_label = QLabel("EMAIL:")
        supplier_email_label.setFont(label_font)
        self.supplier_email_input = QLineEdit()
        self.supplier_email_input.setPlaceholderText("email1@example.com, email2@example.com")
        self.supplier_email_input.setMinimumWidth(250)
        self.supplier_email_input.setStyleSheet("""
            QLineEdit {padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: white;}""")

        left_column.addRow(category_label, self.category_input)
        left_column.addRow(name_label, self.name_input)
        left_column.addRow(description_label, self.description_input)

        right_column.addRow(quantity_label, self.quantity_input)
        right_column.addRow(expiry_label, expiry_layout)
        right_column.addRow(supplier_email_label, self.supplier_email_input)

        form_layout.addLayout(left_column)
        form_layout.addLayout(right_column)
        form_container.setLayout(form_layout)
        main_layout.addWidget(form_container)
        self.product_details_group.setLayout(main_layout)

    # Creates products table.
    def create_table(self):
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "PRODUCT ID", "PRODUCT CATEGORY", "PRODUCT NAME", "DESCRIPTION",
            "QUANTITY", "EMAIL", "BARCODE",
            "TEST DATE", "CREATED DATE", "STATUS"
        ])

        self.table.setStyleSheet("""
            QTableWidget {background-color: white; alternate-background-color: #f7f7f7; gridline-color: #e0e0e0; font-size: 12px;}
            QHeaderView::section {background-color: #f0f0f0; padding: 8px; border: 1px solid #d0d0d0; font-weight: bold;text-align: center;}
            QTableWidget::item {padding: 5px;}
            QTableWidget::item:selected {background-color: #a0c0e0; color: black;}
            QTableWidget::item[status="Approved"] {background-color: #4CAF50; color: white;}
            QTableWidget::item[status="Rejected"] {background-color: #f44336; color: white;}
            QTableWidget::item[status="Pending"] {background-color: #FE9705; color: white;}""")

        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSortingEnabled(True)
        self.table.setWordWrap(True)

        header = self.table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStretchLastSection(True)

        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Category
        header.setSectionResizeMode(2, QHeaderView.Interactive)       # Name
        header.setSectionResizeMode(3, QHeaderView.Interactive)       # Description
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Quantity
        header.setSectionResizeMode(5, QHeaderView.Interactive)       # Supplier Email
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Barcode
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Test Date
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # Created
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)  # Status

        self.table.setColumnWidth(1, 150)  # Category
        self.table.setColumnWidth(2, 200)  # Name
        self.table.setColumnWidth(3, 250)  # Description
        self.table.setColumnWidth(5, 200)  # Supplier Email

        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.table.setTextElideMode(Qt.ElideRight)
        self.table.cellClicked.connect(self.show_product_details)
        self.table.cellDoubleClicked.connect(self.show_barcode_popup)

    # Creates action buttons such Save, Update, Delete and others
    def create_buttons(self):
        self.buttons_container = QWidget()
        container_layout = QVBoxLayout(self.buttons_container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(15)

        button_style = """QPushButton {background-color: #b60338; color: #d9d9d9; border: 1px solid #ccc; border-radius: 4px; padding: 8px 50px; min-width: 100px; font-weight: bold;}
                         QPushButton:hover {background-color: #f31659;}
                         QPushButton:pressed {background-color: #ff4757;}"""

        approve_reject_style = """QPushButton {background-color: #4CAF50; color: white; border: 1px solid #ccc; border-radius: 4px; padding: 8px 50px; min-width: 100px; font-weight: bold;}
                                 QPushButton:hover {background-color: #45a049;}
                                 QPushButton:pressed {background-color: #3e8e41;}"""

        reject_style = """QPushButton {background-color: #f44336; color: white; border: 1px solid #ccc; border-radius: 4px; padding: 8px 50px; min-width: 100px; font-weight: bold;}
                          QPushButton:hover {background-color: #d32f2f;}
                          QPushButton:pressed {background-color: #b71c1c;}"""

        # First row - Approve/Reject buttons
        approval_row = QWidget()
        approval_layout = QHBoxLayout(approval_row)
        approval_layout.setContentsMargins(0, 0, 0, 0)
        approval_layout.addStretch()

        self.approve_button = QPushButton("Approve")
        self.approve_button.setStyleSheet(approve_reject_style)
        self.approve_button.setFixedHeight(35)

        self.reject_button = QPushButton("Reject")
        self.reject_button.setStyleSheet(reject_style)
        self.reject_button.setFixedHeight(35)

        approval_layout.addWidget(self.approve_button)
        approval_layout.addWidget(self.reject_button)
        approval_layout.addStretch()

        first_row_container = QWidget()
        first_row_layout = QHBoxLayout(first_row_container)
        first_row_layout.setContentsMargins(0, 0, 0, 0)
        first_row_layout.addStretch()

        self.save_button = QPushButton("Save")
        self.update_button = QPushButton("Update")
        self.delete_button = QPushButton("Delete")
        self.clear_button = QPushButton("Clear")

        for btn in [self.save_button, self.update_button, self.delete_button, self.clear_button]:
            btn.setStyleSheet(button_style)
            btn.setFixedHeight(35)
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            first_row_layout.addWidget(btn)

        first_row_layout.addStretch()

        second_row_container = QWidget()
        second_row_layout = QHBoxLayout(second_row_container)
        second_row_layout.setContentsMargins(0, 0, 0, 0)
        second_row_layout.addStretch()

        self.export_excel_button = QPushButton("Export to Excel")
        self.export_excel_button.setStyleSheet("""
            QPushButton {background-color: #b60338; color: #d9d9d9; border: 1px solid #ccc; border-radius: 4px; padding: 8px 363px; min-width: 100px; font-weight: bold;}
            QPushButton:hover {background-color: #00ab41;}
            QPushButton:pressed {background-color: #ff4757;}""")

        self.export_excel_button.setFixedHeight(35)
        self.export_excel_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        second_row_layout.addWidget(self.export_excel_button)
        second_row_layout.addStretch()

        container_layout.addWidget(approval_row)
        container_layout.addWidget(first_row_container)
        container_layout.addWidget(second_row_container)

        self.save_button.clicked.connect(self.save_product)
        self.update_button.clicked.connect(self.update_product)
        self.delete_button.clicked.connect(self.delete_product)
        self.clear_button.clicked.connect(self.clear_fields)
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.approve_button.clicked.connect(self.approve_product)
        self.reject_button.clicked.connect(self.reject_product)

    # Adds a row to the table.
    def add_table_row(self, pid, category, name, description, qty, supplier_email="", barcode="", test_date="",
                      created_date=None, status="Pending", creator_username=""):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

        # Format created_date if not provided
        if created_date is None:
            created_date = datetime.now().strftime("%d-%m-%Y")
        elif isinstance(created_date, str):
            try:
                # Handle both database format and display format
                if "-" in created_date and len(created_date.split("-")[0]) == 4:  # Database format (YYYY-MM-DD)
                    created_date = datetime.strptime(created_date, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
            except:
                pass  # Keep original format if conversion fails

        def create_centered_item(text):
            item = QTableWidgetItem(str(text))
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            return item

        def create_left_item(text):
            item = QTableWidgetItem(str(text))
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            return item

        self.table.setItem(row_position, 0, create_centered_item(pid))
        self.table.setItem(row_position, 1, create_left_item(category))
        self.table.setItem(row_position, 2, create_left_item(name))
        self.table.setItem(row_position, 3, create_left_item(description))
        self.table.setItem(row_position, 4, create_centered_item(qty))
        self.table.setItem(row_position, 5, create_left_item(supplier_email))
        self.table.setItem(row_position, 6, create_left_item(barcode))
        self.table.setItem(row_position, 7, create_centered_item(test_date))
        self.table.setItem(row_position, 8, create_centered_item(created_date))
        self.table.setItem(row_position, 9, create_centered_item(status))

        # Status item with color
        status_item = QTableWidgetItem(str(status))
        status_item.setTextAlignment(Qt.AlignCenter)
        status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)

        # Set background color based on status
        if status == "Approved":
            status_item.setBackground(Qt.green)
        elif status == "Rejected":
            status_item.setBackground(Qt.red)
        elif status == "Pending":
            status_item.setBackground(QColor("#FE9705")) 

        self.table.setItem(row_position, 9, status_item)

    def approve_product(self):
        selected_rows = {index.row() for index in self.table.selectedIndexes()}

        if not selected_rows:
            self.show_message("Error", "No rows selected", QMessageBox.Warning)
            return

        # Check if any selected product is already approved or rejected
        for row in sorted(selected_rows):
            status_item = self.table.item(row, 9)
            current_status = status_item.text()
            if current_status in ["Approved", "Rejected"]:
                self.show_message("Error",
                                  f"Cannot change status - product is already {current_status}",
                                  QMessageBox.Warning)
                return

        reply = QMessageBox.question(self, "Confirm Approval",
                                     f"Approve {len(selected_rows)} selected product(s)?",
                                     QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            for row in sorted(selected_rows):
                product_id = int(self.table.item(row, 0).text())
                category = self.table.item(row, 1).text()
                product_name = self.table.item(row, 2).text()
                description = self.table.item(row, 3).text()
                quantity = self.table.item(row, 4).text()
                supplier_email = self.table.item(row, 5).text()
                barcode_path = self.table.item(row, 6).text()
                test_date = self.table.item(row, 7).text()

                try:
                    c = self.db.conn.cursor()
                    c.execute("UPDATE products SET status = 'Approved' WHERE id = ?", (product_id,))
                    self.db.conn.commit()

                    # Update both text and color
                    status_item = self.table.item(row, 9)
                    status_item.setText("Approved")
                    status_item.setBackground(Qt.green)

                    # Send approval email
                    if supplier_email:
                        emails = [e.strip() for e in supplier_email.split(',') if e.strip()]
                        subject = f"Product Approved: {product_name}"
                        body = f"""Dear Product Owner,

Your product "{product_name}" (ID: {product_id}) has been approved.

Product Details:
- Category: {category}
- Description: {description}
- Quantity: {quantity}
- Test Date: {test_date}

Please find the barcode attached for your records.

Thank you,
Shelf Life Management Team"""

                        try:
                            email_sent = self.email_notifier.send_email(emails, subject, body, barcode_path)
                            print("Approval email sent successfully" if email_sent else "Failed to send approval email")
                        except Exception as e:
                            print("Email sending failed with error:", str(e))
                            self.show_message("Email Error", f"Could not send approval email:\n{str(e)}",
                                              QMessageBox.Warning)

                    # Schedule test date reminder emails
                    try:
                        self.schedule_test_date_reminders(product_id, product_name, supplier_email, test_date,
                                                          barcode_path)
                    except Exception as e:
                        print("Reminder scheduling failed:", str(e))


                except Error as e:
                    self.show_message("Error", f"Failed to approve product: {str(e)}", QMessageBox.Critical)

    def reject_product(self):
        selected_rows = {index.row() for index in self.table.selectedIndexes()}

        if not selected_rows:
            self.show_message("Error", "No rows selected", QMessageBox.Warning)
            return

        # Check if any selected product is already approved or rejected
        for row in sorted(selected_rows):
            status_item = self.table.item(row, 9)
            current_status = status_item.text()
            if current_status in ["Approved", "Rejected"]:
                self.show_message("Error",
                                  f"Cannot change status - product is already {current_status}",
                                  QMessageBox.Warning)
                return

        reply = QMessageBox.question(self, "Confirm Rejection",
                                     f"Reject {len(selected_rows)} selected product(s)?",
                                     QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            for row in sorted(selected_rows):
                product_id = int(self.table.item(row, 0).text())
                category = self.table.item(row, 1).text()
                product_name = self.table.item(row, 2).text()
                description = self.table.item(row, 3).text()
                quantity = self.table.item(row, 4).text()
                supplier_email = self.table.item(row, 5).text()
                barcode_path = self.table.item(row, 6).text()
                test_date = self.table.item(row, 7).text()
                try:
                    c = self.db.conn.cursor()
                    c.execute("UPDATE products SET status = 'Rejected' WHERE id = ?", (product_id,))
                    self.db.conn.commit()
                    # Update both text and color
                    status_item = self.table.item(row, 9)
                    status_item.setText("Rejected")
                    status_item.setBackground(Qt.red) 

                    # Send rejection email
                    if supplier_email:
                        emails = [e.strip() for e in supplier_email.split(',') if e.strip()]
                        subject = f"Product Rejected: {product_name}"
                        body = f"""Dear Product Owner,

We regret to inform you that your product "{product_name}" (ID: {product_id}) has been rejected.

Product Details:
- Category: {category}
- Description: {description}
- Quantity: {quantity}
- Test Date: {test_date}

Please find the barcode attached for your records.

If you have any questions, please contact our support team.

Thank you,
Shelf Life Management Team"""

                        if self.email_notifier.send_email(emails, subject, body, barcode_path):
                            print("Rejection email sent successfully")
                        else:
                            print("Failed to send rejection email")

                except Error as e:
                    self.show_message("Error", f"Failed to reject product: {str(e)}", QMessageBox.Critical)

    def schedule_test_date_reminders(self, product_id, product_name, supplier_email, test_date_str, barcode_path=None):
        """Schedule email reminders for test date"""
        if not supplier_email or not test_date_str:
            return

        try:
            test_date = datetime.strptime(test_date_str, "%d-%m-%Y").date()
            today = datetime.now().date()

            emails = [e.strip() for e in supplier_email.split(',') if e.strip()]

            # If test date is today, send immediately
            if test_date == today:
                self.send_test_date_reminder(product_id, product_name, emails, "today", barcode_path)
                return

            # Calculate reminder dates
            two_months_before = test_date - timedelta(days=60)
            one_week_before = test_date - timedelta(days=7)
            one_day_before = test_date - timedelta(days=1)

            # Schedule reminders
            if today <= two_months_before:
                self.schedule_single_reminder(product_id, product_name, emails, two_months_before, "2 months",
                                              barcode_path)

            if today <= one_week_before:
                self.schedule_single_reminder(product_id, product_name, emails, one_week_before, "1 week", barcode_path)

            if today <= one_day_before:
                self.schedule_single_reminder(product_id, product_name, emails, one_day_before, "1 day", barcode_path)

            self.schedule_single_reminder(product_id, product_name, emails, test_date, "today", barcode_path)

        except ValueError:
            print("Invalid test date format")

    def schedule_single_reminder(self, product_id, product_name, emails, reminder_date, time_before, barcode_path=None):
        """Schedule a single reminder email"""
        reminder_time = datetime.combine(reminder_date, datetime.min.time())
        current_time = datetime.now()

        if reminder_time > current_time:
            # Calculate delay in milliseconds
            delay_seconds = (reminder_time - current_time).total_seconds()
            delay_ms = int(delay_seconds * 1000)

            # QTimer only supports 32-bit signed integers (max ~24.8 days in ms)
            if delay_ms > 2_147_483_647:
                print(f"Reminder not scheduled: delay too large ({delay_ms} ms).")
                return  # or store it for future checking
            else:
                timer = QTimer()
                timer.setSingleShot(True)
                timer.timeout.connect(
                    lambda: self.send_test_date_reminder(product_id, product_name, emails, time_before, barcode_path)
                )
                timer.start(delay_ms)
                self.active_timers.append(timer)
        else:
            # If the reminder time has already passed, send immediately
            self.send_test_date_reminder(product_id, product_name, emails, time_before, barcode_path)

    def send_test_date_reminder(self, product_id, product_name, emails, time_before, barcode_path=None):
        """Send test date reminder email"""
        # Get product details from database
        try:
            c = self.db.conn.cursor()
            c.execute("SELECT category, description, quantity, test_date FROM products WHERE id = ?", (product_id,))
            product_data = c.fetchone()

            if product_data:
                category, description, quantity, test_date = product_data
                if time_before == "today":
                    subject = f"Test Date Today: {product_name}"
                    body = f"""Dear Product Owner,

This is to notify you that today is the test date for product "{product_name}" (ID: {product_id}).

Product Details:
- Category: {category}
- Description: {description}
- Quantity: {quantity}
- Test Date: {test_date}

Please ensure all necessary tests are conducted as scheduled and attached barcode for your record. 

Thank you,
Shelf Life Management Team"""

                else:
                    subject = f"Test Date Reminder: {product_name} ({time_before} before)"
                    body = f"""Dear Product Owner,

This is a reminder that the test date for product "{product_name}" (ID: {product_id}) 
is coming up in {time_before}.

Product Details:
- Category: {category}
- Description: {description}
- Quantity: {quantity}
- Test Date: {test_date}

Please make necessary arrangements and attached barcode for your record.

Thank you,
Shelf Life Management Team"""

                self.email_notifier.send_email(emails, subject, body, barcode_path)
        except Error as e:
            print(f"Error sending reminder: {e}")

    # Saves a new product to database.
    def save_product(self):
        try:
            category = self.category_input.text().strip()
            name = self.name_input.text().strip()
            description = self.description_input.toPlainText().strip()
            qty = self.quantity_input.text().strip()
            test_date = self.expiry_date_input.text().strip()
            supplier_email = self.supplier_email_input.text().strip()

            if not category:
                self.show_message("Error", "Please enter a category", QMessageBox.Warning)
                return

            if not name:
                self.show_message("Error", "Product name cannot be empty", QMessageBox.Warning)
                return

            if not qty or not qty.isdigit():
                self.show_message("Error", "Please enter a valid quantity", QMessageBox.Warning)
                return

            if not self.validate_date(test_date):
                return  # validate_date will show appropriate error message

            if not supplier_email:
                self.show_message("Error", "email is required", QMessageBox.Warning)
                return

            if supplier_email and not self.validate_email(supplier_email):
                self.show_message("Error",
                                  "Please enter valid email\n"
                                  "Multiple emails should be comma-separated\n"
                                  "Example: supplier1@example.com, supplier2@domain.com",
                                  QMessageBox.Warning)
                return

            for row in range(self.table.rowCount()):
                existing_name = self.table.item(row, 2).text()
                if existing_name.lower() == name.lower():
                    reply = QMessageBox.question(self, "Confirm",
                                                 f"A product with name '{name}' already exists.\nDo you want to create a new entry with a different barcode?",
                                                 QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.No:
                        return
                    break

            barcode_number, filename = self.barcode_gen.generate_barcode(name)

            product = (category, name, description, int(qty), supplier_email, filename,
                       test_date if test_date else None)
            product_id = self.db.add_product(product)

            if product_id:
                self.add_table_row(product_id, category, name, description, qty, supplier_email, filename, test_date)
                self.clear_fields()
                self.show_message("Success", "Product saved successfully!")
            else:
                self.show_message("Error", "Failed to save product to database", QMessageBox.Critical)

        except Exception as e:
            self.show_message("Error", f"An error occurred: {str(e)}", QMessageBox.Critical)

    # Updates an existing product.
    def update_product(self):
        try:
            selected = self.table.selectedItems()
            if not selected:
                self.show_message("Error", "Please select a row to update", QMessageBox.Warning)
                return

            row = selected[0].row()
            status_item = self.table.item(row, 9) 

            # Check if product status is not "Pending"
            if status_item.text() != "Pending":
                self.show_message("Error", "Can only update products with 'Pending' status", QMessageBox.Warning)
                return

            reply = QMessageBox.question(self, "Confirm Update", "Are you sure you want to update this product?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply != QMessageBox.Yes:
                return

            row = selected[0].row()
            product_id = int(self.table.item(row, 0).text())
            old_name = self.table.item(row, 2).text()  
            old_barcode_path = self.table.item(row, 6).text()  
            category = self.category_input.text().strip()
            new_name = self.name_input.text().strip()
            description = self.description_input.toPlainText().strip()
            qty = self.quantity_input.text().strip()
            test_date = self.expiry_date_input.text().strip()
            supplier_email = self.supplier_email_input.text().strip()

            if not category:
                self.show_message("Error", "Please enter a category", QMessageBox.Warning)
                return

            if not new_name:
                self.show_message("Error", "Product name cannot be empty", QMessageBox.Warning)
                return

            if not qty or not qty.isdigit():
                self.show_message("Error", "Please enter a valid quantity", QMessageBox.Warning)
                return

            if test_date and not self.validate_date(test_date):
                self.show_message("Error", "Please enter date in DD-MM-YYYY format (e.g. 30-05-2025)",
                                  QMessageBox.Warning)
                return

            if supplier_email and not self.validate_email(supplier_email):
                self.show_message("Error",
                                  "Please enter a valid email address\n"
                                  "Example: supplier@example.com",
                                  QMessageBox.Warning)
                return

            # Handle barcode file renaming if product name changed
            new_barcode_path = old_barcode_path
            if new_name != old_name and old_barcode_path and os.path.exists(old_barcode_path):
                try:
                    # Get directory and barcode number from old path
                    dir_name = os.path.dirname(old_barcode_path)
                    barcode_num = os.path.basename(old_barcode_path).split('_')[-1].split('.')[0]

                    # Create new safe filename
                    safe_new_name = "".join(c if c.isalnum() else "_" for c in new_name)
                    new_filename = f"{safe_new_name}_{barcode_num}.png"
                    new_barcode_path = os.path.join(dir_name, new_filename)

                    # Rename the files
                    os.rename(old_barcode_path, new_barcode_path)

                    # Also rename the .pnm file if it exists
                    pnm_old = old_barcode_path.replace('.png', '.pnm')
                    pnm_new = new_barcode_path.replace('.png', '.pnm')
                    if os.path.exists(pnm_old):
                        os.rename(pnm_old, pnm_new)

                except Exception as e:
                    self.show_message("Warning",
                                      f"Product updated but could not rename barcode file: {str(e)}",
                                      QMessageBox.Warning)
                    new_barcode_path = old_barcode_path  # Keep old path if rename failed

            product = (category, new_name, description, int(qty), supplier_email if supplier_email else None,
                       new_barcode_path, test_date if test_date else None, product_id)

            updated_rows = self.db.update_product(product)

            if updated_rows:
                self.table.item(row, 1).setText(category)
                self.table.item(row, 2).setText(new_name)
                self.table.item(row, 3).setText(description)
                self.table.item(row, 4).setText(qty)
                self.table.item(row, 5).setText(supplier_email)
                self.table.item(row, 6).setText(new_barcode_path) 
                self.table.item(row, 7).setText(test_date)

                self.clear_fields()
                self.show_message("Success", "Product updated successfully!")
            else:
                self.show_message("Error", "Failed to update product in database", QMessageBox.Critical)

        except Exception as e:
            self.show_message("Error", f"An error occurred: {str(e)}", QMessageBox.Critical)

    # Deletes selected products.
    def delete_product(self):
        selected_rows = {index.row() for index in self.table.selectedIndexes()}

        if not selected_rows:
            self.show_message("Error", "No rows selected", QMessageBox.Warning)
            return

        # Check if any selected product is pending approval
        for row in selected_rows:
            status_item = self.table.item(row, 9) 
            if status_item and status_item.text() == "Pending":
                self.show_message("Error",
                                  "Cannot delete products with 'Pending' status. Approve or reject first.",
                                  QMessageBox.Warning)
                return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete {len(selected_rows)} selected product(s)?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            for row in sorted(selected_rows, reverse=True):
                product_id = int(self.table.item(row, 0).text())
                barcode_item = self.table.item(row, 6) 

                deleted_rows = self.db.delete_product(product_id)

                if deleted_rows:
                    if barcode_item and barcode_item.text():
                        try:
                            filepath = barcode_item.text()
                            if os.path.exists(filepath):
                                os.remove(filepath)
                            # Remove the .pnm file if it exists
                            pnm_path = filepath.replace('.png', '.pnm')
                            if os.path.exists(pnm_path):
                                os.remove(pnm_path)
                        except Exception as e:
                            self.show_message("Warning",
                                              f"Deleted product but couldn't delete image: {str(e)}",
                                              QMessageBox.Warning)

                    self.table.removeRow(row)
                else:
                    self.show_message("Error", f"Failed to delete product ID {product_id}", QMessageBox.Critical)

    # Clears the product form.
    def clear_fields(self):
        self.category_input.clear()
        self.name_input.clear()
        self.description_input.clear()
        self.quantity_input.clear()
        self.expiry_date_input.clear()
        self.supplier_email_input.clear()

    # Searches products based on criteria.
    def search_products(self):
        search_type = self.search_combo.currentText()

        if search_type == "Select":
            self.show_message("Error", "Please select search type", QMessageBox.Warning)
            return

        search_term = self.search_input.text().strip()

        # For status searches, we don't need a search term
        if not search_term and not search_type.startswith("Status:"):
            self.show_message("Error", "Please enter search term", QMessageBox.Warning)
            return

        # Clear current filters
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)

        if search_type.startswith("Status:"):
            status = search_type.split(":")[1].strip()
            for row in range(self.table.rowCount()):
                status_item = self.table.item(row, 9) 
                if status_item and status_item.text() != status:
                    self.table.setRowHidden(row, True)
        else:
            products = self.db.search_products(search_type, search_term)

            # If searching by something other than date, filter the table
            if search_type != "Date Range":
                for row in range(self.table.rowCount()):
                    match = False
                    if search_type == "ID":
                        item = self.table.item(row, 0)
                        if item and search_term == item.text():
                            match = True
                    elif search_type == "Name":
                        item = self.table.item(row, 2)
                        if item and search_term.lower() in item.text().lower():
                            match = True
                    elif search_type == "Category":
                        item = self.table.item(row, 1)
                        if item and search_term.lower() in item.text().lower():
                            match = True
                    elif search_type == "Description":
                        item = self.table.item(row, 3)
                        if item and search_term.lower() in item.text().lower():
                            match = True

                    self.table.setRowHidden(row, not match)

    # Shows product details when row is clicked.
    def show_product_details(self, row, column):
        try:
            category = self.table.item(row, 1).text()
            name = self.table.item(row, 2).text()
            description = self.table.item(row, 3).text()
            qty = self.table.item(row, 4).text()
            supplier_email = self.table.item(row, 5).text()
            test_date = self.table.item(row, 7).text()  

            self.category_input.setText(category)
            self.name_input.setText(name)
            self.description_input.setPlainText(description)
            self.quantity_input.setText(qty)
            self.expiry_date_input.setText(test_date)
            self.supplier_email_input.setText(supplier_email)

        except Exception as e:
            self.show_message("Error", f"Error loading product details: {str(e)}", QMessageBox.Critical)

    # Shows barcode popup when barcode cell is double-clicked.
    def show_barcode_popup(self, row, column):
        if column == 6:  
            barcode_item = self.table.item(row, column)
            if barcode_item and barcode_item.text():
                barcode_path = barcode_item.text()
                self.popup = BarcodePopup(barcode_path, self)
                self.popup.show()

    # Exports table data to Excel with barcode images.
    def export_to_excel(self):
        if self.table.rowCount() == 0:
            self.show_message("Error", "No data to export", QMessageBox.Warning)
            return

        try:
            file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")

            if not file_path:
                return

            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'

            headers = []
            for col in range(self.table.columnCount()):
                headers.append(self.table.horizontalHeaderItem(col).text())

            data = []
            for row in range(self.table.rowCount()):
                if not self.table.isRowHidden(row):  # Only export visible rows
                    row_data = []
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)

            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)

            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Products')

                workbook = writer.book
                worksheet = writer.sheets['Products']

                # Set column widths
                worksheet.column_dimensions['A'].width = 12  # ID
                worksheet.column_dimensions['B'].width = 20  # Category
                worksheet.column_dimensions['C'].width = 30  # Name
                worksheet.column_dimensions['D'].width = 40  # Description
                worksheet.column_dimensions['E'].width = 10  # Quantity
                worksheet.column_dimensions['F'].width = 25  # Supplier Email
                worksheet.column_dimensions['G'].width = 30  # Barcode
                worksheet.column_dimensions['H'].width = 12  # Test Date
                worksheet.column_dimensions['I'].width = 12  # Created
                worksheet.column_dimensions['J'].width = 12  # Status

                from openpyxl.styles import PatternFill, Font, Alignment
                wrap_alignment = Alignment(wrap_text=True, vertical='top')
                from openpyxl.styles.colors import Color

                # Define the fill colors
                approved_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                rejected_fill = PatternFill(start_color='F44336', end_color='F44336', fill_type='solid')
                pending_fill = PatternFill(start_color='FE9705', end_color='FE9705', fill_type='solid')

                # Define white font for better visibility
                white_font = Font(color='FFFFFF')

                # Apply status colors
                for row_idx, row_data in enumerate(data, start=2):
                    status_cell = worksheet[f'J{row_idx}'] 
                    status = status_cell.value

                    if status == "Approved":
                        status_cell.fill = approved_fill
                        status_cell.font = white_font
                    elif status == "Rejected":
                        status_cell.fill = rejected_fill
                        status_cell.font = white_font
                    elif status == "Pending":
                        status_cell.fill = pending_fill
                        status_cell.font = white_font

                for row_idx in range(2, len(data) + 2):  # Start from row 2 skip header
                    worksheet.row_dimensions[row_idx].height = 60  # Default height

                    # Apply wrap text to description column D and email column F
                    worksheet[f'D{row_idx}'].alignment = wrap_alignment
                    worksheet[f'F{row_idx}'].alignment = wrap_alignment

                    # Auto-adjust row height based on content length
                    desc_len = len(str(worksheet[f'D{row_idx}'].value))
                    email_len = len(str(worksheet[f'F{row_idx}'].value))

                    # Increase row height if content is long
                    if desc_len > 100 or email_len > 50:
                        worksheet.row_dimensions[row_idx].height = 80
                    if desc_len > 200 or email_len > 100:
                        worksheet.row_dimensions[row_idx].height = 100

                # Create center alignment style for all cells
                center_alignment = Alignment(horizontal='center', vertical='center')

                # Apply center alignment to all cells
                for row in worksheet.iter_rows(min_row=2, max_row=len(data) + 1):
                    for cell in row:
                        # Center align most columns
                        if cell.column_letter not in ['D', 'F']:
                            cell.alignment = center_alignment

                # Add images to cells with centered alignment
                for row_idx, row_data in enumerate(data, start=2):  # Skip header
                    barcode_path = row_data[6] 
                    if barcode_path and os.path.exists(barcode_path):
                        try:
                            img = Image(barcode_path)

                            # Resize image to fit cell
                            img.width = 120
                            img.height = 50

                            # Calculate cell coordinates
                            cell = f'G{row_idx}'

                            # Create an anchor point for the image to center it
                            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
                            from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
                            from openpyxl.utils.units import pixels_to_EMU

                            # Calculate center position
                            col_width = worksheet.column_dimensions['G'].width
                            row_height = worksheet.row_dimensions[row_idx].height or 15

                            # Convert dimensions to EMU (Excel Measurement Units)
                            cell_width_emu = pixels_to_EMU(col_width * 7)  
                            cell_height_emu = pixels_to_EMU(row_height)
                            img_width_emu = pixels_to_EMU(img.width)
                            img_height_emu = pixels_to_EMU(img.height)

                            # Calculate offset to center the image
                            col_offset = (cell_width_emu - img_width_emu) / 2
                            row_offset = (cell_height_emu - img_height_emu) / 2

                            # Create anchor with centered position
                            img.anchor = OneCellAnchor(
                                _from=AnchorMarker(
                                    col=6,  
                                    colOff=col_offset,
                                    row=row_idx - 1,  
                                    rowOff=row_offset
                                ),
                                ext=XDRPositiveSize2D(
                                    img_width_emu,
                                    img_height_emu
                                )
                            )

                            # Add image to worksheet
                            worksheet.add_image(img)

                            # Clear the text (we're showing the image)
                            worksheet[cell] = None

                        except Exception as e:
                            print(f"Error adding barcode image: {e}")
                            # Keep the filename if image fails to load
                            worksheet[f'G{row_idx}'] = barcode_path
                    else:
                        # If no barcode path, ensure cell is empty
                        worksheet[f'G{row_idx}'] = ""

            self.show_message("Success", f"Filtered data exported to Excel successfully!\n{file_path}")

        except Exception as e:
            self.show_message("Error", f"Export failed: {str(e)}", QMessageBox.Critical)

    # Resets table to show all products.
    def show_all_products(self):
        """Show all rows and clear search"""
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
        self.search_input.clear()
        self.search_combo.setCurrentIndex(0)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ProductManager()
    window.show()
    sys.exit(app.exec_())
