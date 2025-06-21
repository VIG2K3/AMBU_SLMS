import sys
import sqlite3
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QPushButton, QLineEdit, QLabel, QMessageBox,
    QComboBox, QFileDialog, QInputDialog, QHeaderView, QCheckBox,
    QGroupBox, QFormLayout, QSplitter, QSizePolicy, QStyle
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QIcon
import hashlib
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import csv


# Hashing function
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# Database setup
def create_employee_table():
    conn = sqlite3.connect("employees.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            email TEXT,
            role TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

# Default admin
def default_admin_Cred():
    conn = sqlite3.connect("employees.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM employees WHERE username = ?", ("admin",))
    if not cursor.fetchone():
        default_password = hash_password("admin123")
        cursor.execute("""
            INSERT INTO employees (first_name, last_name, username, password, email, role)
            VALUES (?, ?, ?, ?, ?, ?)
        """, ("Enter First Name", "Enter Last Name", "admin", default_password, "admin@example.com", "Admin"))
        conn.commit()
    conn.close()

# Calls correct order
create_employee_table()
default_admin_Cred()


class ProductOwnerRegistration(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Employee Registration System")
        self.resize(1200, 800)
        self.selected_emp_id = None
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        self.setLayout(main_layout)

        # Search Group (top section)
        self.create_search_group()
        main_layout.addWidget(self.search_group)

        # Create splitter for table and employee details
        self.splitter = QSplitter(Qt.Vertical)
        main_layout.addWidget(self.splitter)

        # Create table widget
        self.create_table_widget()
        table_container = QWidget()
        table_container.setLayout(QVBoxLayout())
        table_container.layout().addWidget(self.table)
        table_container.layout().setContentsMargins(0, 0, 0, 0)
        self.splitter.addWidget(table_container)

        # Create employee details form
        self.create_employee_form()
        form_container = QWidget()
        form_container.setLayout(QVBoxLayout())
        form_container.layout().addWidget(self.employee_details_group)
        form_container.layout().setContentsMargins(0, 0, 0, 0)
        self.splitter.addWidget(form_container)

        # Set initial sizes (table takes 70%, form takes 30%)
        self.splitter.setSizes([int(self.height() * 0.7), int(self.height() * 0.3)])

        # Create buttons (bottom section)
        self.create_buttons()
        main_layout.addWidget(self.buttons_container)

        # Apply styles
        self.setStyleSheet("""
            QWidget { font-family: 'Segoe UI';}
            QGroupBox {font-size: 14px; font-weight: bold; border: 1px solid #ccc; border-radius: 5px; margin-top: 10px; padding-top: 15px;}
            QGroupBox::title {subcontrol-origin: margin; left: 10px; padding: 0 3px;}
            QTableWidget {background-color: white; alternate-background-color: #f7f7f7; gridline-color: #e0e0e0; font-size: 12px;}
            QHeaderView::section {background-color: #f0f0f0; padding: 8px; border: 1px solid #d0d0d0; font-weight: bold; text-align: center;}
            QTableWidget::item {padding: 5px;}
            QTableWidget::item:selected {background-color: #a0c0e0; color: black;}
            QPushButton {background-color: #b60338; color: #d9d9d9; border: 1px solid #ccc; border-radius: 4px; padding: 8px 50px; min-width: 100px; font-weight: bold;}
            QPushButton:hover { background-color: #f31659;}
            QPushButton:pressed { background-color: #ff4757;}
            QLineEdit, QComboBox {padding: 5px; border: 1px solid #ccc; border-radius: 3px;}
            QCheckBox { spacing: 5px;}""")

    def create_search_group(self):
        self.search_group = QGroupBox("SEARCH BAR")
        self.search_group.setStyleSheet("""
            QGroupBox {font-size: 14px; font-weight: bold; border: 1px solid #ccc; border-radius: 5px; margin-top: 10px; padding-top: 15px;}
            QGroupBox::title {subcontrol-origin: margin; left: 10px; padding: 0 3px;}""")

        vbox = QVBoxLayout()
        vbox.setContentsMargins(20, 20, 20, 20)

        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(15)

        search_layout.addStretch()

        self.search_combo = QComboBox()
        self.search_combo.addItems(["Select", "ID", "First Name", "Last Name", "Username", "Email", "Role"])
        self.search_combo.setStyleSheet("""
            QComboBox {padding: 8px; border: 1px solid #ccc; border-radius: 3px; min-width: 100px; background-color: white;}""")
        search_layout.addWidget(self.search_combo)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search...")
        self.search_input.setStyleSheet("""
            QLineEdit {padding: 8px; border: 1px solid #ccc; border-radius: 3px; min-width: 300px; background-color: white;}""")
        search_layout.addWidget(self.search_input)

        self.search_button = QPushButton("Search")
        self.search_button.setStyleSheet("""
            QPushButton {background-color: #b60338; color: white; border: 1px solid #ccc; padding: 8px; font-weight: bold; min-width: 100px;}
            QPushButton:hover {background-color: #f31659;}
            QPushButton:pressed {background-color: #ff4757;}""")
        search_layout.addWidget(self.search_button)

        self.show_all_button = QPushButton("Show All")
        self.show_all_button.setStyleSheet("""
            QPushButton {background-color: #b60338; color: white; border: 1px solid #ccc; padding: 8px; font-weight: bold; min-width: 100px;}
            QPushButton:hover {background-color: #f31659;}
            QPushButton:pressed {background-color: #ff4757;}""")

        search_layout.addWidget(self.show_all_button)
        search_layout.addStretch()

        vbox.addWidget(search_container)
        self.search_group.setLayout(vbox)

        self.search_button.clicked.connect(self.search_employees)
        self.show_all_button.clicked.connect(self.load_data)

    def create_table_widget(self):
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "ID", "FIRST NAME", "LAST NAME",
            "USERNAME", "PASSWORD", "EMAIL", "ROLE"
        ])

        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSortingEnabled(True)
        self.table.setWordWrap(True)

        header = self.table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStretchLastSection(True)

        # Set all columns to have equal width using Stretch
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.Stretch)

        self.table.setTextElideMode(Qt.ElideRight)
        self.table.cellClicked.connect(self.load_selected_row)

    def create_employee_form(self):
        self.employee_details_group = QGroupBox("REGISTER PRODUCT OWNER")
        self.employee_details_group.setStyleSheet("""
            QGroupBox {font-size: 14px; font-weight: bold; border: 1px solid #ccc; border-radius: 5px; margin-top: 10px; padding-top: 15px;}
            QGroupBox::title {subcontrol-origin: margin; left: 10px; padding: 0 3px;}""")

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignCenter)
        main_layout.setContentsMargins(20, 20, 20, 20)

        form_container = QWidget()
        form_container.setMaximumWidth(900)
        form_layout = QHBoxLayout()
        form_layout.setContentsMargins(20, 10, 20, 10)
        form_layout.setSpacing(30)

        left_column = QFormLayout()
        left_column.setFormAlignment(Qt.AlignLeft)
        left_column.setLabelAlignment(Qt.AlignLeft)
        left_column.setSpacing(15)
        left_column.setContentsMargins(10, 10, 10, 10)

        right_column = QFormLayout()
        right_column.setFormAlignment(Qt.AlignLeft)
        right_column.setLabelAlignment(Qt.AlignLeft)
        right_column.setSpacing(15)
        right_column.setContentsMargins(10, 10, 10, 10)

        label_font = QFont()
        label_font.setBold(True)

        # Left column fields
        first_name_label = QLabel("FIRST NAME:")
        first_name_label.setFont(label_font)
        self.first_name_input = QLineEdit()
        self.first_name_input.setPlaceholderText("Enter first name")
        self.first_name_input.setStyleSheet("background-color: white;")
        self.first_name_input.setMinimumWidth(250)

        last_name_label = QLabel("LAST NAME:")
        last_name_label.setFont(label_font)
        self.last_name_input = QLineEdit()
        self.last_name_input.setPlaceholderText("Enter last name")
        self.last_name_input.setStyleSheet("background-color: white;")
        self.last_name_input.setMinimumWidth(250)

        email_label = QLabel("EMAIL:")
        email_label.setFont(label_font)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("user@example.com")
        self.email_input.setStyleSheet("background-color: white;")
        self.email_input.setMinimumWidth(250)

        left_column.addRow(first_name_label, self.first_name_input)
        left_column.addRow(last_name_label, self.last_name_input)
        left_column.addRow(email_label, self.email_input)

        # Right column fields
        username_label = QLabel("USERNAME:")
        username_label.setFont(label_font)
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Enter username")
        self.username_input.setStyleSheet("background-color: white;")
        self.username_input.setMinimumWidth(250)

        password_label = QLabel("PASSWORD:")
        password_label.setFont(label_font)
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Enter password")
        self.password_input.setStyleSheet("background-color: white;")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setMinimumWidth(250)

        self.toggle_password_checkbox = QCheckBox("View Password")
        self.toggle_password_checkbox.toggled.connect(self.toggle_password_visibility)

        # Fixed Role
        role_label = QLabel("ROLE:")
        role_label.setFont(label_font)
        self.role_display = QLabel("Product Owner")
        self.role_display.setStyleSheet("background-color: white; padding: 7px; border-radius: 3px; border: 1px solid #ccc;")
        right_column.addRow(role_label, self.role_display)

        # Right Position
        right_column.addRow(username_label, self.username_input)
        right_column.addRow(password_label, self.password_input)
        right_column.addRow(QLabel(""), self.toggle_password_checkbox)

        form_layout.addLayout(left_column)
        form_layout.addLayout(right_column)
        form_container.setLayout(form_layout)
        main_layout.addWidget(form_container)
        self.employee_details_group.setLayout(main_layout)

    def create_buttons(self):
        self.buttons_container = QWidget()
        container_layout = QVBoxLayout(self.buttons_container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(15)

        button_style = """QPushButton {background-color: #b60338; color: white; border: 1px solid #ccc; border-radius: 4px; padding: 8px 50px; min-width: 100px; font-weight: bold; font-family: Segoe UI;}
                         QPushButton:hover {background-color: #f31659;}
                         QPushButton:pressed {background-color: #ff4757;}"""

        first_row_container = QWidget()
        first_row_layout = QHBoxLayout(first_row_container)
        first_row_layout.setContentsMargins(0, 0, 0, 0)
        first_row_layout.addStretch()

        self.add_button = QPushButton("Save")
        self.update_button = QPushButton("Update")
        self.delete_button = QPushButton("Delete")
        self.clear_button = QPushButton("Clear")

        for btn in [self.add_button, self.update_button, self.delete_button, self.clear_button]:
            btn.setStyleSheet(button_style)
            btn.setFixedHeight(35)
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            first_row_layout.addWidget(btn)

        first_row_layout.addStretch()

        second_row_container = QWidget()
        second_row_layout = QHBoxLayout(second_row_container)
        second_row_layout.setContentsMargins(0, 0, 0, 0)
        second_row_layout.addStretch()

        self.export_button = QPushButton("Export To Excel")
        self.export_button.setStyleSheet("""
            QPushButton {background-color: #b60338; color: white; border: 1px solid #ccc; border-radius: 4px; padding: 8px 363px; min-width: 100px; font-weight: bold; font-family: Segoe UI;}
            QPushButton:hover {background-color: #f31659;}
            QPushButton:pressed {background-color: #ff4757;}""")

        self.export_button.setFixedHeight(35)
        self.export_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        second_row_layout.addWidget(self.export_button)
        second_row_layout.addStretch()

        container_layout.addWidget(first_row_container)
        container_layout.addWidget(second_row_container)

        self.add_button.clicked.connect(self.add_employee)
        self.update_button.clicked.connect(self.update_employee)
        self.delete_button.clicked.connect(self.delete_employee)
        self.clear_button.clicked.connect(self.clear_fields)
        self.export_button.clicked.connect(self.export_to_excel)

    def toggle_password_visibility(self, checked):
        self.password_input.setEchoMode(QLineEdit.Normal if checked else QLineEdit.Password)

    def load_data(self):
        self.table.setRowCount(0)
        conn = sqlite3.connect("employees.db")
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM employees
            ORDER BY CASE WHEN username = 'admin' THEN 0 ELSE 1 END, ID
        """)

        for row_data in cursor.fetchall():
            row = self.table.rowCount()
            self.table.insertRow(row)

            for col, data in enumerate(row_data):
                item = QTableWidgetItem(str(data))
                if col == 4:  # Password column - show asterisks
                    item.setText("********")
                item.setTextAlignment(Qt.AlignCenter if col != 1 and col != 2 else Qt.AlignLeft | Qt.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row, col, item)
        conn.close()

    def load_selected_row(self, row, column):
        self.selected_emp_id = self.table.item(row, 0).text()
        self.first_name_input.setText(self.table.item(row, 1).text())
        self.last_name_input.setText(self.table.item(row, 2).text())
        self.username_input.setText(self.table.item(row, 3).text())
        self.password_input.clear()
        self.toggle_password_checkbox.setChecked(False)
        self.email_input.setText(self.table.item(row, 5).text())

        role = self.table.item(row, 6).text()  # Or adjust index if needed
        self.role_display.setText(role)

    def search_employees(self):
        search_type = self.search_combo.currentText()
        search_term = self.search_input.text().strip()

        if search_type == "Select":
            self.show_message("Error", "Please select search type", QMessageBox.Warning)
            return

        if not search_term:
            self.show_message("Error", "Please enter search term", QMessageBox.Warning)
            return

        self.table.setRowCount(0)
        conn = sqlite3.connect("employees.db")
        cursor = conn.cursor()

        if search_type == "ID":
            cursor.execute("SELECT * FROM employees WHERE empID = ?", (search_term,))
        elif search_type == "First Name":
            cursor.execute("SELECT * FROM employees WHERE first_name LIKE ?", (f'%{search_term}%',))
        elif search_type == "Last Name":
            cursor.execute("SELECT * FROM employees WHERE last_name LIKE ?", (f'%{search_term}%',))
        elif search_type == "Username":
            cursor.execute("SELECT * FROM employees WHERE username LIKE ?", (f'%{search_term}%',))
        elif search_type == "Email":
            cursor.execute("SELECT * FROM employees WHERE email LIKE ?", (f'%{search_term}%',))
        elif search_type == "Role":
            cursor.execute("SELECT * FROM employees WHERE role LIKE ?", (f'%{search_term}%',))

        for row_data in cursor.fetchall():
            row = self.table.rowCount()
            self.table.insertRow(row)

            for col, data in enumerate(row_data):
                item = QTableWidgetItem(str(data))
                if col == 4:  # Password column - show asterisks
                    item.setText("********")
                item.setTextAlignment(Qt.AlignCenter if col != 1 and col != 2 else Qt.AlignLeft | Qt.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row, col, item)

        conn.close()

    def show_message(self, title, message, icon=QMessageBox.Information):
        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setIcon(icon)
        msg.exec_()

    def add_employee(self):
        first = self.first_name_input.text().strip()
        last = self.last_name_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        email = self.email_input.text().strip()
        role = "Product Owner"


        if not all([first, last, username, password, email]):
            self.show_message("Input Error", "All fields are required.", QMessageBox.Warning)
            return

        # Email format validation
        email_pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if not re.match(email_pattern, email):
            self.show_message("Invalid Email", "Please enter a valid email address (e.g., example@mail.com).",
                              QMessageBox.Warning)
            return

        if len(password) < 8:
            self.show_message("Input Error", "Password must be at least 8 characters long.", QMessageBox.Warning)
            return

        hashed = hash_password(password)
        try:
            conn = sqlite3.connect("employees.db")
            cursor = conn.cursor()

            # Check for duplicate email
            cursor.execute("SELECT * FROM employees WHERE email = ?", (email,))
            if cursor.fetchone():
                self.show_message("Duplicate Email", "This email is already in use.", QMessageBox.Warning)
                conn.close()
                return

            # Check for duplicate password
            cursor.execute("SELECT * FROM employees WHERE password = ?", (hashed,))
            if cursor.fetchone():
                self.show_message("Duplicate Password",
                                  "This password is already in use by another employee. Choose a different password.",
                                  QMessageBox.Warning)
                conn.close()
                return

            confirm = QMessageBox.question(self, "Add Employee", "Are you sure you want to add this employee?",
                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if confirm != QMessageBox.Yes:
                conn.close()
                return

            cursor.execute(
                "INSERT INTO employees (first_name, last_name, username, password, email, role) VALUES (?, ?, ?, ?, ?, ?)",
                (first, last, username, hashed, email, role))
            conn.commit()
            conn.close()
            self.load_data()
            self.clear_fields()
        except sqlite3.IntegrityError:
            self.show_message("Error", "Username already exists.", QMessageBox.Warning)

    def update_employee(self):
        if not self.selected_emp_id:
            self.show_message("Selection Error", "Please select a row to update.", QMessageBox.Warning)
            return

        first = self.first_name_input.text().strip()
        last = self.last_name_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        email = self.email_input.text().strip()

        if not all([first, last, username, email]):
            self.show_message("Input Error", "All fields except password are required.", QMessageBox.Warning)
            return

        # Lock role based on username
        if username == "admin":
            role = "Admin"
        else:
            role = "Product Owner"

        conn = sqlite3.connect("employees.db")
        cursor = conn.cursor()

        confirm = QMessageBox.question(self, "Update Employee", "Are you sure you want to update this employee?",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if confirm != QMessageBox.Yes:
            conn.close()
            return

        if password:
            if len(password) < 8:
                self.show_message("Input Error", "Password must be at least 8 characters long.", QMessageBox.Warning)
            else:
                self.show_message("Password Update Not Allowed",
                                  "You cannot update the password after account creation.", QMessageBox.Warning)
            conn.close()
            return  # Block update if password field is touched

        # Proceed with updating other fields only
        cursor.execute(
            "UPDATE employees SET first_name=?, last_name=?, username=?, email=?, role=? WHERE ID=?",
            (first, last, username, email, role, self.selected_emp_id)
        )

        conn.commit()
        conn.close()
        self.load_data()
        self.clear_fields()

    def delete_employee(self):
        if not self.selected_emp_id:
            self.show_message("Selection Error", "Please select a row to delete.", QMessageBox.Warning)
            return

        username = self.username_input.text().strip()
        if username == "admin":
            self.show_message("Permission Denied", "The default admin account cannot be deleted.", QMessageBox.Critical)
            return

        conn = sqlite3.connect("employees.db")
        cursor = conn.cursor()

        confirm = QMessageBox.question(
            self, "Confirm Deletion", "Are you sure you want to delete this employee?",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            conn.close()
            return

        cursor.execute("DELETE FROM employees WHERE ID = ?", (self.selected_emp_id,))
        conn.commit()
        conn.close()

        self.load_data()
        self.clear_fields()

    def export_to_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "employees.xlsx", "Excel Files (*.xlsx)")
        if path:
            try:
                conn = sqlite3.connect("employees.db")
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM employees")
                records = cursor.fetchall()
                conn.close()

                wb = Workbook()
                ws = wb.active
                ws.title = "Employees"

                headers = ["ID", "First Name", "Last Name", "Username", "Password", "Email", "Role"]
                ws.append(headers)

                # Make headers bold
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)

                # Write data
                for row in records:
                    ws.append(row)

                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column].width = adjusted_width

                wb.save(path)
                self.show_message("Export Successful", f"Data exported to {path}")
            except Exception as e:
                self.show_message("Export Failed", f"An error occurred: {str(e)}", QMessageBox.Critical)

    def clear_fields(self):
        self.first_name_input.clear()
        self.last_name_input.clear()
        self.username_input.clear()
        self.password_input.clear()
        self.email_input.clear()
        self.toggle_password_checkbox.setChecked(False)
        self.role_display.setText("Product Owner")
        self.selected_emp_id = None


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ProductOwnerRegistration()
    window.show()
    sys.exit(app.exec_())